import os
import glob
from openpyxl import load_workbook
from ebooklib import epub

def create_epub_from_xlsx(xlsx_file):
    # Extract book title and author from file name
    file_name = os.path.splitext(os.path.basename(xlsx_file))[0]
    parts = file_name.split('_')
    if len(parts) >= 2:
        book_title = '_'.join(parts[:-1])
        author = parts[-1]
    else:
        book_title = file_name
        author = "Unknown"

    # Create a new EPUB book
    book = epub.EpubBook()

    # Set metadata
    book.set_identifier(f'id_{book_title}')
    book.set_title(book_title)
    book.set_language('zh-TW')
    book.add_author(author)  # Add author to metadata

    # Load the XLSX file
    wb = load_workbook(xlsx_file)
    ws = wb.active

    chapters = []
    for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):
        if row[0] and row[1]:  # Ensure both title and content are present
            chapter_title, chapter_content = row

            # Replace newline characters with HTML line breaks
            formatted_content = chapter_content.replace('\n', '<br>')

            chapter = epub.EpubHtml(title=chapter_title, file_name=f'chap_{len(chapters)+1}.xhtml', lang='zh-TW')
            chapter.content = f'<h1>{chapter_title}</h1><p>{formatted_content}</p>'
            book.add_item(chapter)
            chapters.append(chapter)

    # Define Table of Contents
    book.toc = chapters

    # Add default NCX and Nav file
    book.add_item(epub.EpubNcx())
    book.add_item(epub.EpubNav())

    # Define CSS style
    style = '''
    @namespace epub "http://www.idpf.org/2007/ops";
    body {
        font-family: Arial, sans-serif;
    }
    h1 {
        text-align: center;
        color: #333333;
    }
    p {
        text-indent: 1em;
        margin-bottom: 0.5em;
    }
    '''

    nav_css = epub.EpubItem(uid="style_nav", file_name="style/nav.css", media_type="text/css", content=style)
    book.add_item(nav_css)

    # Define spine
    book.spine = ['nav'] + chapters

    # Create EPUB file
    epub_filename = f'{book_title} - {author}.epub'
    epub.write_epub(epub_filename, book, {})

    print(f"EPUB file '{epub_filename}' has been created successfully.")

# Process all XLSX files in the current directory
xlsx_files = glob.glob('*.xlsx')

if xlsx_files:
    for xlsx_file in xlsx_files:
        create_epub_from_xlsx(xlsx_file)
    print(f"Processed {len(xlsx_files)} XLSX files.")
else:
    print("No XLSX files found in the current directory.")