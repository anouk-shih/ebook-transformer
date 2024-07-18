import os
import glob
from openpyxl import load_workbook
from ebooklib import epub
import requests
from PIL import Image, UnidentifiedImageError
from io import BytesIO

def create_comic_epub(xlsx_file):
    # Extract comic title and author from file name
    file_name = os.path.splitext(os.path.basename(xlsx_file))[0]
    parts = file_name.split('_')
    if len(parts) >= 2:
        comic_title = '_'.join(parts[:-1])
        author = parts[-1]
    else:
        comic_title = file_name
        author = "Unknown"

    # Create a new EPUB book
    book = epub.EpubBook()

    # Set metadata
    book.set_identifier(f'id_{comic_title}')
    book.set_title(comic_title)
    book.set_language('zh-TW')  # Set language to Chinese
    book.add_author(author)  # Add author to metadata

    # Load the XLSX file
    wb = load_workbook(xlsx_file)
    ws = wb.active

    # Get the key (column name) from the first row
    key = ws[1][0].value

    pages = []
    for index, row in enumerate(ws.iter_rows(min_row=2, max_col=1, values_only=True), start=1):
        if row[0]:  # Ensure the cell is not empty
            image_url = row[0]

            try:
                # Download the image
                response = requests.get(image_url, timeout=10)
                response.raise_for_status()  # Raise an exception for bad status codes

                # Try to open the image
                try:
                    img = Image.open(BytesIO(response.content))
                except UnidentifiedImageError:
                    print(f"Skipping unidentified image at row {index}: {image_url}")
                    continue

                # Convert image to JPEG if it's not already
                if img.format != 'JPEG':
                    img_byte_arr = BytesIO()
                    img.convert('RGB').save(img_byte_arr, format='JPEG')
                    img_data = img_byte_arr.getvalue()
                else:
                    img_data = response.content

                # Create a unique filename for the image
                image_filename = f'image_{len(pages)+1}.jpg'

                # Add the image to the EPUB
                epub_image = epub.EpubImage(uid=image_filename, file_name=f'Images/{image_filename}', media_type='image/jpeg', content=img_data)
                book.add_item(epub_image)

                # Create a chapter for the image
                chapter = epub.EpubHtml(title=f'Page {len(pages)+1}', file_name=f'page_{len(pages)+1}.xhtml', lang='zh')
                chapter.content = f'<img src="Images/{image_filename}" alt="Page {len(pages)+1}"/>'
                book.add_item(chapter)
                pages.append(chapter)

            except requests.RequestException as e:
                print(f"Error downloading image at row {index}: {e}")
            except Exception as e:
                print(f"Error processing image at row {index}: {e}")

    # Define Table of Contents
    book.toc = pages

    # Add default NCX and Nav file
    book.add_item(epub.EpubNcx())
    book.add_item(epub.EpubNav())

    # Define CSS style
    style = '''
    @namespace epub "http://www.idpf.org/2007/ops";
    body {
        margin: 0;
        padding: 0;
    }
    img {
        max-width: 100%;
        height: auto;
    }
    '''

    nav_css = epub.EpubItem(uid="style_nav", file_name="style/nav.css", media_type="text/css", content=style)
    book.add_item(nav_css)

    # Define spine
    book.spine = ['nav'] + pages

    # Create EPUB file
    epub_filename = f'{comic_title}.epub'
    epub.write_epub(epub_filename, book, {})

    print(f"Comic EPUB file '{epub_filename}' has been created successfully.")

# Process all XLSX files in the 'comic' folder
script_dir = os.path.dirname(os.path.abspath(__file__))
comic_folder = os.path.join(script_dir, 'comic')
xlsx_files = glob.glob(os.path.join(comic_folder, '*.xlsx'))

if xlsx_files:
    for xlsx_file in xlsx_files:
        create_comic_epub(xlsx_file)
    print(f"Processed {len(xlsx_files)} comic XLSX files.")
else:
    print("No XLSX files found in the 'comic' folder.")